"""
CEREBRO PHASE 1 — HARM TOLERANCE CLOCK
Data Ingestion Pipeline
======================
Sources pulled automatically:
  - FRED API: Unemployment, civilian employment, incarceration proxy
  - World Bank API: Homicide rate
  - BJS CSV: Incarceration rate 1925-2022
  - CDC WONDER: Drug overdose deaths (manual CSV required — see instructions)

Velocity & Acceleration: Computed automatically after data assembly.
Saddle Point Detection: Fires when inflection conditions met.

Output:
  - cerebro_harm_clock_phase1.xlsx  (data + calculations + saddle flags)
  - cerebro_harm_clock_data.csv     (raw for reuse)

Author: Cerebro Project / Jennifer Leigh West / The Forgotten Code Research Institute
"""

import requests
import pandas as pd
import json
import time
import os
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────

FRED_API_KEY = os.environ.get("FRED_API_KEY", "")
FRED_BASE = "https://api.stlouisfed.org/fred/series/observations"
WB_BASE = "https://api.worldbank.org/v2"

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
YEARS_START = 1960
YEARS_END = 2024

print("=" * 60)
print("CEREBRO PHASE 1: HARM TOLERANCE CLOCK — DATA INGESTION")
print(f"Pulling {YEARS_START}–{YEARS_END}")
print("=" * 60)

# ─────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────

def fred_series(series_id, api_key=FRED_API_KEY, start=str(YEARS_START), end=str(YEARS_END)):
    """Pull annual data from FRED. Returns dict {year: value}."""
    if not api_key:
        print(f"  ✗ FRED {series_id} failed: FRED_API_KEY not set")
        return {}
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "observation_start": f"{start}-01-01",
        "observation_end": f"{end}-12-31",
        "frequency": "a",           # annual aggregation
        "aggregation_method": "avg"
    }
    try:
        r = requests.get(FRED_BASE, params=params, timeout=15)
        data = r.json()
        result = {}
        for obs in data.get("observations", []):
            year = int(obs["date"][:4])
            try:
                result[year] = float(obs["value"])
            except ValueError:
                pass  # "." = missing
        print(f"  ✓ FRED {series_id}: {len(result)} observations")
        return result
    except Exception as e:
        print(f"  ✗ FRED {series_id} failed: {e}")
        return {}

def worldbank_series(indicator, country="US", start=YEARS_START, end=YEARS_END):
    """Pull annual data from World Bank API. Returns dict {year: value}."""
    url = f"{WB_BASE}/country/{country}/indicator/{indicator}"
    params = {
        "format": "json",
        "date": f"{start}:{end}",
        "per_page": 200
    }
    try:
        r = requests.get(url, params=params, timeout=15)
        data = r.json()
        if len(data) < 2:
            return {}
        result = {}
        for obs in data[1]:
            if obs["value"] is not None:
                result[int(obs["date"])] = float(obs["value"])
        print(f"  ✓ World Bank {indicator}: {len(result)} observations")
        return result
    except Exception as e:
        print(f"  ✗ World Bank {indicator} failed: {e}")
        return {}

def compute_velocity(series, window=3):
    """Rolling N-year rate of change. Returns dict {year: velocity}."""
    years = sorted(series.keys())
    velocity = {}
    for i, yr in enumerate(years):
        if i >= window:
            prev_yr = years[i - window]
            if prev_yr in series and series[prev_yr] != 0:
                velocity[yr] = (series[yr] - series[prev_yr]) / window
    return velocity

def compute_acceleration(velocity, window=3):
    """Second derivative of position = rate of change in velocity."""
    years = sorted(velocity.keys())
    acceleration = {}
    for i, yr in enumerate(years):
        if i >= window:
            prev_yr = years[i - window]
            if prev_yr in velocity:
                acceleration[yr] = (velocity[yr] - velocity[prev_yr]) / window
    return acceleration

def detect_saddle(velocity, acceleration):
    """
    Saddle signal: velocity < 0 AND acceleration > 0 AND
    acceleration crossed zero from below in last 3 years.
    Returns dict {year: signal_strength (0-3)}
    """
    years = sorted(set(velocity.keys()) & set(acceleration.keys()))
    saddle = {}
    for i, yr in enumerate(years):
        score = 0
        v = velocity.get(yr, 0)
        a = acceleration.get(yr, 0)
        if v < 0:      score += 1   # still moving negative
        if a > 0:      score += 1   # decelerating (momentum loading)
        # Check if acceleration recently crossed zero
        for lookback in [1, 2, 3]:
            if i >= lookback:
                prev_yr = years[i - lookback]
                prev_a = acceleration.get(prev_yr, 0)
                if prev_a < 0 and a > 0:
                    score += 1
                    break
        saddle[yr] = score
    return saddle

def normalize(series, target_min=-1, target_max=1):
    """Normalize series to [-1, +1] range using historical min/max."""
    if not series:
        return {}
    vals = list(series.values())
    lo, hi = min(vals), max(vals)
    if hi == lo:
        return {yr: 0 for yr in series}
    return {yr: target_min + (v - lo) / (hi - lo) * (target_max - target_min)
            for yr, v in series.items()}


# ─────────────────────────────────────────
# SECTION 1: PULL LIVE DATA
# ─────────────────────────────────────────

print("\n[1/5] FRED ECONOMIC SERIES")

# Unemployment rate — leads crime/harm clock 6-12 months
unemployment = fred_series("UNRATE")

# Labor force participation — structural harm signal
lfpr = fred_series("CIVPART")

# Real median household income — class/harm intersection
median_income = fred_series("MEHOINUSA672N")

# Personal saving rate — social safety buffer signal
saving_rate = fred_series("PSAVERT")

# Federal minimum wage (annual avg) — labor harm structural
# Use Employment Cost Index as proxy since min wage needs manual
eci = fred_series("ECIALLCIV")

time.sleep(1)

print("\n[2/5] WORLD BANK SERIES")

# Intentional homicide rate per 100k — Ring C behavioral
homicide_wb = worldbank_series("VC.IHR.PSRC.P5", country="US")

# Prison population — use FRED BJS proxy
# FRED doesn't have direct incarceration series, use crime index
# We'll embed the BJS historical data manually below

time.sleep(1)

print("\n[3/5] EMBEDDING HISTORICAL BJS INCARCERATION DATA")
# Source: BJS National Prisoner Statistics 1925-2022
# https://bjs.ojp.gov/data-collection/national-prisoner-statistics-nps-program
# US incarceration rate per 100,000 US residents (sentenced prisoners in state/federal)
# Manually compiled from BJS reports — this is the authoritative series

bjs_incarceration = {
    1960: 117, 1961: 119, 1962: 115, 1963: 114, 1964: 111,
    1965: 108, 1966: 102, 1967: 98,  1968: 94,  1969: 97,
    1970: 96,  1971: 95,  1972: 93,  1973: 96,  1974: 102,
    1975: 111, 1976: 120, 1977: 129, 1978: 132, 1979: 133,
    1980: 139, 1981: 154, 1982: 170, 1983: 178, 1984: 188,
    1985: 202, 1986: 216, 1987: 231, 1988: 247, 1989: 276,
    1990: 297, 1991: 313, 1992: 332, 1993: 359, 1994: 389,
    1995: 411, 1996: 427, 1997: 444, 1998: 461, 1999: 476,
    2000: 478, 2001: 470, 2002: 476, 2003: 482, 2004: 487,
    2005: 491, 2006: 501, 2007: 506, 2008: 504, 2009: 502,
    2010: 500, 2011: 492, 2012: 480, 2013: 478, 2014: 471,
    2015: 459, 2016: 450, 2017: 440, 2018: 431, 2019: 419,
    2020: 358, 2021: 336, 2022: 350
}
# Filter to years range
bjs_incarceration = {yr: v for yr, v in bjs_incarceration.items()
                     if YEARS_START <= yr <= YEARS_END}
print(f"  ✓ BJS Incarceration: {len(bjs_incarceration)} observations (embedded)")

print("\n[4/5] EMBEDDING HISTORICAL FBI HOMICIDE DATA")
# Source: FBI UCR Supplemental Homicide Reports 1960-2022
# Rate per 100,000 population
fbi_homicide = {
    1960: 5.1, 1961: 4.8, 1962: 4.6, 1963: 4.6, 1964: 4.9,
    1965: 5.1, 1966: 5.6, 1967: 6.2, 1968: 6.9, 1969: 7.3,
    1970: 7.9, 1971: 8.6, 1972: 9.0, 1973: 9.4, 1974: 9.8,
    1975: 9.6, 1976: 8.7, 1977: 8.8, 1978: 9.0, 1979: 9.7,
    1980: 10.2,1981: 9.8, 1982: 9.1, 1983: 8.3, 1984: 7.9,
    1985: 8.0, 1986: 8.6, 1987: 8.3, 1988: 8.5, 1989: 8.7,
    1990: 9.4, 1991: 9.8, 1992: 9.3, 1993: 9.5, 1994: 9.0,
    1995: 8.2, 1996: 7.4, 1997: 6.8, 1998: 6.3, 1999: 5.7,
    2000: 5.5, 2001: 5.6, 2002: 5.6, 2003: 5.7, 2004: 5.5,
    2005: 5.6, 2006: 5.8, 2007: 5.7, 2008: 5.4, 2009: 5.0,
    2010: 4.8, 2011: 4.7, 2012: 4.7, 2013: 4.5, 2014: 4.4,
    2015: 4.9, 2016: 5.4, 2017: 5.3, 2018: 5.0, 2019: 5.0,
    2020: 6.5, 2021: 6.9, 2022: 6.3
}
fbi_homicide = {yr: v for yr, v in fbi_homicide.items()
                if YEARS_START <= yr <= YEARS_END}
print(f"  ✓ FBI Homicide Rate: {len(fbi_homicide)} observations (embedded)")

print("\n[5/5] EMBEDDING DRUG OVERDOSE DEATH RATE")
# Source: CDC WONDER, NCHS Drug Poisoning Mortality 1968-2022
# Rate per 100,000 population (age-adjusted)
# Manual pull from: https://www.cdc.gov/nchs/nvss/vsrr/drug-overdose-data.htm
cdc_overdose = {
    1968: 1.6, 1969: 1.7, 1970: 1.9, 1971: 2.2, 1972: 2.3,
    1973: 2.1, 1974: 2.2, 1975: 2.2, 1976: 2.1, 1977: 2.2,
    1978: 2.2, 1979: 2.4, 1980: 2.5, 1981: 2.3, 1982: 2.3,
    1983: 2.3, 1984: 2.4, 1985: 2.5, 1986: 2.8, 1987: 2.9,
    1988: 3.1, 1989: 3.4, 1990: 3.4, 1991: 3.5, 1992: 3.5,
    1993: 3.8, 1994: 3.9, 1995: 4.0, 1996: 4.4, 1997: 4.6,
    1998: 4.5, 1999: 4.6, 2000: 4.9, 2001: 5.5, 2002: 6.9,
    2003: 7.5, 2004: 7.9, 2005: 8.8, 2006: 10.0,2007: 10.4,
    2008: 11.0,2009: 11.6,2010: 12.3,2011: 13.0,2012: 13.1,
    2013: 13.8,2014: 14.7,2015: 16.3,2016: 19.8,2017: 21.7,
    2018: 20.7,2019: 21.6,2020: 27.1,2021: 32.4,2022: 32.0
}
cdc_overdose = {yr: v for yr, v in cdc_overdose.items()
                if YEARS_START <= yr <= YEARS_END}
print(f"  ✓ CDC Drug Overdose Rate: {len(cdc_overdose)} observations (embedded)")


# ─────────────────────────────────────────
# SECTION 2: ASSEMBLE MASTER DATAFRAME
# ─────────────────────────────────────────

print("\n[ASSEMBLING] Building master dataframe...")

years = list(range(YEARS_START, min(YEARS_END + 1, 2025)))
df = pd.DataFrame(index=years)
df.index.name = "year"

# Raw indicator columns
df["unemployment_rate"]       = df.index.map(unemployment)
df["labor_force_part_rate"]   = df.index.map(lfpr)
df["median_household_income"] = df.index.map(median_income)
df["saving_rate"]             = df.index.map(saving_rate)
df["homicide_rate_fbi"]       = df.index.map(fbi_homicide)
df["homicide_rate_wb"]        = df.index.map(homicide_wb)
df["incarceration_rate_bjs"]  = df.index.map(bjs_incarceration)
df["overdose_death_rate_cdc"] = df.index.map(cdc_overdose)

# Prefer FBI homicide, fill with World Bank where missing
df["homicide_rate"] = df["homicide_rate_fbi"].fillna(df["homicide_rate_wb"])

# Leading indicators (2–4 year lead on saddle points)
DATA_DIR = os.path.join(OUTPUT_DIR, "cerebro_data")
li_path = os.path.join(DATA_DIR, "WorldBank_leading_indicators_US.csv")
acled_path = os.path.join(DATA_DIR, "ACLED_protest_annual.csv")
if os.path.exists(li_path):
    try:
        li_df = pd.read_csv(li_path)
        if "year" in li_df.columns:
            for col in ["youth_unemployment_pct", "tertiary_enrollment_pct", "birth_rate_per_1000"]:
                if col in li_df.columns:
                    d = dict(zip(li_df["year"].astype(int), li_df[col]))
                    df[col] = df.index.map(d)
        print("  ✓ Leading indicators: youth_unemp, tertiary_enrollment, birth_rate")
    except Exception:
        pass
# UCDP/ACLED protest: disabled — conflict counts ≠ protest intensity, contaminated signal
# if os.path.exists(acled_path):
#     try:
#         acled_df = pd.read_csv(acled_path)
#         ...


# ─────────────────────────────────────────
# SECTION 3: RING SCORING
# ─────────────────────────────────────────

print("[SCORING] Computing ring positions...")

"""
HARM TOLERANCE CLOCK RING ASSIGNMENTS

Ring C (Behavioral) — 30% weight — what people actually DO
  C1: Homicide rate (higher = more harm, negative pole)
  C2: Drug overdose death rate (higher = more harm)
  C3: Incarceration rate (higher = more punitive, positive pole — society
      is *responding* to harm with enforcement; at high levels it *becomes* harm)

Ring B (Normative) — 30% weight — what attitudes say
  B1: Gallup tough-on-crime approval [MANUAL — placeholder 0.5 for now]
  B2: Trust in government Pew [MANUAL — placeholder]

Ring A (Structural) — 40% weight — laws and institutions
  A1: Incarceration rate proxy for criminalization (also behavioral but structural signal)
  A2: Unemployment rate proxy for structural economic harm loading
  A3: Labor force participation inverse (lower LFPR = more structural exclusion)

NOTE: Ring B requires manual GSS/Gallup data. Placeholders marked.
      Run with real data once CSVs downloaded from GSS Data Explorer.
"""

# NORMALIZE each indicator to [-1, +1] — CAUSAL (expanding min/max, no future leakage)
# Convention: -1 = maximum harm / constraint pole; +1 = maximum safety/protection pole

from cerebro_causal_normalization import norm_causal

# Ring C — Behavioral
# Higher homicide = more harm = negative pole → invert
rc1_homicide    = norm_causal(dict(zip(df.index, df["homicide_rate"])), invert=True)
# Higher overdose = more harm = negative pole → invert
rc2_overdose    = norm_causal(dict(zip(df.index, df["overdose_death_rate_cdc"])), invert=True)
# Incarceration: dual signal. Very high = structural harm. For Phase 1, treat rising as HARM.
rc3_incarcerate = norm_causal(dict(zip(df.index, df["incarceration_rate_bjs"])), invert=True)

# Ring B — Normative: GSS attitudes + Pew trust
# Load GSS Ring B if parsed (cerebro_data/GSS_RingB_annual.csv)
gss_path = os.path.join(OUTPUT_DIR, "cerebro_data", "GSS_RingB_annual.csv")
pew_path = os.path.join(OUTPUT_DIR, "cerebro_data", "PEW_trust_government.csv")
pew_trust = {}
gss_ringb = None

if os.path.exists(pew_path):
    try:
        pew_df = pd.read_csv(pew_path)
        if "year" in pew_df.columns and "trust_pct" in pew_df.columns:
            pew_trust = dict(zip(pew_df["year"].astype(int), pew_df["trust_pct"]))
    except Exception:
        pass

if os.path.exists(gss_path):
    try:
        gss_df = pd.read_csv(gss_path, index_col=0)
        gss_df.index = gss_df.index.astype(int)
        # Harm Tolerance Ring B composite: punitive (inv) + trust (pos) + fear (inv)
        # courts_not_harsh_enough, death_penalty_favor, police_violence_ok: higher = punitive = negative
        # trust_people: higher = positive. afraid_night: higher = fear = negative
        components = []
        for col, inv in [
            ("courts_not_harsh_enough_pct", True),
            ("death_penalty_favor_pct", True),
            ("police_violence_ok_pct", True),
            ("trust_people_pct", False),
            ("afraid_night_pct", True),
        ]:
            if col in gss_df.columns:
                s = gss_df[col].dropna()
                if len(s) > 5:
                    d = s.to_dict()
                    components.append(norm_causal(d, invert=inv))
        if components:
            rb_gss = sum(components) / len(components)
            gss_ringb = rb_gss.reindex(years)
    except Exception as e:
        pass

# Prefer GSS composite; fall back to Pew trust; else placeholder
if gss_ringb is not None and gss_ringb.notna().sum() > 10:
    df["ring_B_score"] = gss_ringb
    # Fill pre-1972 with Pew if available
    if pew_trust:
        pew_series = norm_causal(pew_trust, invert=False).reindex(years)
        df["ring_B_score"] = df["ring_B_score"].fillna(pew_series)
    print("  ✓ Ring B: GSS attitudes (COURTS, CAPPUN, TRUST, FEAR, etc.) + Pew trust")
elif pew_trust:
    rb_pew = norm_causal(pew_trust, invert=False)
    df["ring_B_score"] = rb_pew.reindex(years)
    print("  ✓ Ring B: Pew trust in government (1958–present)")
else:
    df["ring_B_score"] = pd.Series(0.0, index=years)
    print("  ⚠ Ring B: Placeholder — run cerebro_gss_loader.py + cerebro_data_gather.py")

# Ring A — Structural
# Higher unemployment = more harm loading = negative pole → invert
ra1_unemployment = norm_causal(dict(zip(df.index, df["unemployment_rate"])), invert=True)
# Lower LFPR = more exclusion = harm = negative → invert
ra2_lfpr         = norm_causal(dict(zip(df.index, df["labor_force_part_rate"])), invert=False)
# Incarceration rate also as structural signal
ra3_incarcerate_struct = norm_causal(dict(zip(df.index, df["incarceration_rate_bjs"])), invert=True)

# FRED available? Check if unemployment pulled successfully
fred_available = ra1_unemployment.notna().sum() > 10

if fred_available:
    df["ring_A_score"] = (ra1_unemployment + ra2_lfpr + ra3_incarcerate_struct) / 3
    ring_a_note = "Ring A: Unemployment + LFPR + Incarceration"
else:
    # FRED unavailable (no API key) — use incarceration only for Ring A
    df["ring_A_score"] = ra3_incarcerate_struct
    ring_a_note = "Ring A: Incarceration proxy only (add FRED key for full A ring)"
    print(f"  ⚠ FRED data unavailable. Add real API key at top of script.")
    print(f"    Get free key at: https://fred.stlouisfed.org/docs/api/api_key.html")
    print(f"    Ring A computed from BJS incarceration only for now.")

# Ring C — available regardless (embedded)
df["ring_C_score"] = (rc1_homicide + rc2_overdose + rc3_incarcerate) / 3
# Ring B set above (Pew trust or placeholder)

# Leading component (signals precede saddle by 2–4 years)
leading_components = []
if "youth_unemployment_pct" in df.columns and df["youth_unemployment_pct"].notna().sum() > 10:
    r_lead_youth = norm_causal(dict(zip(df.index, df["youth_unemployment_pct"])), invert=True)
    leading_components.append(r_lead_youth.reindex(years))
if "tertiary_enrollment_pct" in df.columns and df["tertiary_enrollment_pct"].notna().sum() > 10:
    r_lead_tert = norm_causal(dict(zip(df.index, df["tertiary_enrollment_pct"])), invert=False)
    leading_components.append(r_lead_tert.reindex(years))
if "birth_rate_per_1000" in df.columns and df["birth_rate_per_1000"].notna().sum() > 10:
    r_lead_birth = norm_causal(dict(zip(df.index, df["birth_rate_per_1000"])), invert=True)
    leading_components.append(r_lead_birth.reindex(years))
# UCDP protest: disabled (zero weight)
if leading_components:
    df["leading_score"] = pd.concat(leading_components, axis=1).mean(axis=1)
    leading_available = True
    names = []
    if "youth_unemployment_pct" in df.columns and df["youth_unemployment_pct"].notna().sum() > 10:
        names.append("youth_unemp")
    if "tertiary_enrollment_pct" in df.columns and df["tertiary_enrollment_pct"].notna().sum() > 10:
        names.append("tertiary")
    if "birth_rate_per_1000" in df.columns and df["birth_rate_per_1000"].notna().sum() > 10:
        names.append("birth_rate")
    # UCDP protest: disabled
    print("  ✓ Leading component:", ", ".join(names) if names else "none")
else:
    df["leading_score"] = pd.Series(0.0, index=years)
    leading_available = False

# CLOCK POSITION SCORE
# Leading indicators get 15% when available (path to Brier 0.077)
if fred_available and leading_available:
    df["clock_position"] = (
        0.35 * df["ring_A_score"] +
        0.25 * df["ring_B_score"] +
        0.25 * df["ring_C_score"] +
        0.15 * df["leading_score"]
    )
elif fred_available:
    df["clock_position"] = (
        0.40 * df["ring_A_score"] +
        0.30 * df["ring_B_score"] +
        0.30 * df["ring_C_score"]
    )
else:
    # Without FRED, weight Ring A less, Ring C more
    w_lead = 0.15 if leading_available else 0
    df["clock_position"] = (
        (0.30 - w_lead * 0.5) * df["ring_A_score"] +
        0.20 * df["ring_B_score"] +
        (0.50 - w_lead * 0.5) * df["ring_C_score"]
    )
    if leading_available:
        df["clock_position"] = df["clock_position"] + w_lead * df["leading_score"]
    print(f"  Weights adjusted: Ring A 30%, Ring B 20%, Ring C 50% (add FRED for full weighting)")

# Scale to -10 to +10 for readability
df["clock_position_10pt"] = df["clock_position"] * 10


# ─────────────────────────────────────────
# SECTION 4: VELOCITY & ACCELERATION
# ─────────────────────────────────────────

print("[DYNAMICS] Computing velocity and acceleration...")

pos_dict = df["clock_position"].dropna().to_dict()

velocity_dict     = compute_velocity(pos_dict, window=3)
# L1 Google Trends: blend cultural velocity when available (leading 3–12 months)
trends_path = os.path.join(DATA_DIR, "GoogleTrends_cultural_velocity.csv")
if os.path.exists(trends_path):
    try:
        trends_df = pd.read_csv(trends_path, index_col=0, parse_dates=True)
        if "cultural_velocity_smooth" in trends_df.columns and "year" in trends_df.columns:
            vy = trends_df.groupby("year")["cultural_velocity_smooth"].mean()
            for yr in df.index:
                if yr in velocity_dict and yr in vy.index:
                    v_computed = velocity_dict[yr]
                    v_trends = float(vy.loc[yr]) / 100.0
                    velocity_dict[yr] = 0.85 * v_computed + 0.15 * v_trends
            print("  ✓ L1 Google Trends cultural velocity blended into harm clock")
    except Exception:
        pass
acceleration_dict = compute_acceleration(velocity_dict, window=3)
saddle_dict       = detect_saddle(velocity_dict, acceleration_dict)

df["velocity"]     = df.index.map(velocity_dict)
df["acceleration"] = df.index.map(acceleration_dict)
df["saddle_score"] = df.index.map(saddle_dict)

# Saddle label
def saddle_label(score):
    if pd.isna(score):  return ""
    score = int(score)
    if score == 3:      return "STRONG SADDLE ⚡"
    elif score == 2:    return "MODERATE SADDLE"
    elif score == 1:    return "WEAK SIGNAL"
    return ""
df["saddle_label"] = df["saddle_score"].apply(saddle_label)

# Composite Pressure Index (single clock version)
df["pressure_index"] = df["clock_position"] * df["velocity"].abs()


# ─────────────────────────────────────────
# SECTION 5: EXPORT TO CSV
# ─────────────────────────────────────────

csv_path = os.path.join(OUTPUT_DIR, "cerebro_harm_clock_data.csv")
df.to_csv(csv_path)
print(f"\n[EXPORT] CSV saved: {csv_path}")


# ─────────────────────────────────────────
# SECTION 6: BUILD EXCEL WORKBOOK
# ─────────────────────────────────────────

print("[EXPORT] Building Excel workbook...")
xlsx_path = os.path.join(OUTPUT_DIR, "cerebro_harm_clock_phase1.xlsx")

try:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Color palette ──
    BLACK    = "FF000000"
    WHITE    = "FFFFFFFF"
    DARK_BG  = "FF0D0D1A"   # near-black
    MID_BG   = "FF1A1A2E"
    ACCENT   = "FF6C63FF"   # purple
    WARN_RED = "FFFF4040"
    SAFE_GRN = "FF40FF80"
    GOLD     = "FFFFD700"
    BLUE     = "FF4080FF"
    GRAY     = "FF888888"
    LITE_BG  = "FFF5F5F5"
    SADDLE_YELLOW = "FFFFFF00"
    SADDLE_ORANGE = "FFFF8C00"
    SADDLE_RED    = "FFFF2200"

    def hdr_font(bold=True, size=11, color=WHITE):
        return Font(name="Arial", bold=bold, size=size, color=color)

    def cell_font(bold=False, size=10, color=BLACK):
        return Font(name="Arial", bold=bold, size=size, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="FFCCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def right_align():
        return Alignment(horizontal="right", vertical="center")

    # ════════════════════════════════════════
    # SHEET 1: README / METHODOLOGY
    # ════════════════════════════════════════
    ws_readme = wb.active
    ws_readme.title = "📋 README"
    ws_readme.sheet_view.showGridLines = False

    ws_readme.column_dimensions["A"].width = 3
    ws_readme.column_dimensions["B"].width = 28
    ws_readme.column_dimensions["C"].width = 70

    # Title block
    ws_readme.row_dimensions[1].height = 10
    ws_readme.row_dimensions[2].height = 40
    ws_readme.merge_cells("B2:C2")
    c = ws_readme["B2"]
    c.value = "CEREBRO — HARM TOLERANCE CLOCK / PHASE 1"
    c.font = Font(name="Arial", bold=True, size=18, color=WHITE)
    c.fill = fill(DARK_BG)
    c.alignment = center()

    ws_readme.row_dimensions[3].height = 22
    ws_readme.merge_cells("B3:C3")
    c = ws_readme["B3"]
    c.value = "Pendulum Dynamics Engine | The Forgotten Code Research Institute"
    c.font = Font(name="Arial", size=11, color=GRAY)
    c.fill = fill(MID_BG)
    c.alignment = center()

    rows = [
        ("", ""),
        ("WHAT THIS IS", "A predictive engine measuring the Harm Tolerance Clock — one of four clocks in the Cerebro system. Harm Tolerance tracks how much harm a society allows its members to inflict on each other and themselves, across three dimensions: structural (laws, institutions), normative (attitudes), and behavioral (what people actually do)."),
        ("", ""),
        ("THE PHYSICS", "At the nadir of a pendulum's swing, velocity is maximum. 'Rock bottom' is a static metaphor — this engine uses a kinetic model. The predictive signal is not 'how bad is it' but 'is the downward acceleration decelerating?' That inflection = the Saddle Point = maximum momentum loading for reversal."),
        ("", ""),
        ("THREE LAYERS", "POSITION: Where is the clock on its axis right now? (-10 = maximum harm; +10 = maximum protection)\nVELOCITY: How fast is it changing, and in what direction?\nACCELERATION: Is velocity speeding up or slowing down? When deceleration hits zero = SADDLE SIGNAL."),
        ("", ""),
        ("DATA SOURCES", "C1 Homicide Rate: FBI UCR 1960-2022 (embedded)\nC2 Drug Overdose Deaths: CDC NCHS 1968-2022 (embedded)\nC3 Incarceration Rate: BJS National Prisoner Statistics 1960-2022 (embedded)\nA1 Unemployment Rate: FRED UNRATE (live pull)\nA2 Labor Force Participation: FRED CIVPART (live pull)\nA3 Median Household Income: FRED MEHOINUSA672N (live pull)\nB RING: ⚠ PLACEHOLDER — load GSS/Gallup CSVs to populate"),
        ("", ""),
        ("SADDLE SIGNAL", "Score 1 = Weak signal\nScore 2 = Moderate (sector-level inflection possible)\nScore 3 = STRONG ⚡ (era-level inflection — watch leading indicators now)"),
        ("", ""),
        ("NEXT STEPS", "1. Download GSS data for Ring B attitudes (gssdataexplorer.norc.org)\n2. Download Gallup tough-on-crime series\n3. Add GDELT media tone velocity column\n4. Validate saddle detections against known historical turning points\n5. Replicate for other 3 clocks\n6. Build Composite Pressure Index across all clocks"),
        ("", ""),
        ("TABS IN THIS FILE", "📋 README — This page\n📊 Raw Data — All raw indicator series by year\n⚙ Clock Engine — Position, Velocity, Acceleration, Saddle scores\n📈 Charts — Visual clock history\n🔮 Saddle Signals — Saddle detection events only"),
    ]

    for i, (label, text) in enumerate(rows, start=4):
        r = 3 + i
        ws_readme.row_dimensions[r].height = max(20, len(text.split("\n")) * 16)
        b_cell = ws_readme.cell(row=r, column=2)
        c_cell = ws_readme.cell(row=r, column=3)
        b_cell.value = label
        c_cell.value = text
        if label:
            b_cell.font = Font(name="Arial", bold=True, size=10, color=ACCENT)
            b_cell.fill = fill("FF16213E")
            c_cell.fill = fill("FF16213E")
        else:
            c_cell.fill = fill(MID_BG)
        c_cell.font = Font(name="Arial", size=10, color=WHITE)
        c_cell.alignment = Alignment(wrap_text=True, vertical="top")
        b_cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ════════════════════════════════════════
    # SHEET 2: RAW DATA
    # ════════════════════════════════════════
    ws_raw = wb.create_sheet("📊 Raw Data")
    ws_raw.sheet_view.showGridLines = False

    raw_cols = [
        ("year", "Year"),
        ("unemployment_rate", "Unemployment Rate (%)"),
        ("labor_force_part_rate", "Labor Force Participation (%)"),
        ("median_household_income", "Median HH Income (2022$)"),
        ("saving_rate", "Personal Saving Rate (%)"),
        ("homicide_rate", "Homicide Rate (per 100k)"),
        ("incarceration_rate_bjs", "Incarceration Rate (per 100k)"),
        ("overdose_death_rate_cdc", "OD Death Rate (per 100k)"),
    ]

    col_widths = [8, 20, 25, 25, 22, 22, 25, 25]
    for i, w in enumerate(col_widths, 1):
        ws_raw.column_dimensions[get_column_letter(i)].width = w

    ws_raw.row_dimensions[1].height = 10
    ws_raw.row_dimensions[2].height = 14
    ws_raw.row_dimensions[3].height = 36

    # Header row
    for ci, (field, label) in enumerate(raw_cols, 1):
        c = ws_raw.cell(row=3, column=ci)
        c.value = label
        c.font = hdr_font(size=10)
        c.fill = fill(DARK_BG)
        c.alignment = center()
        c.border = thin_border()

    # Data rows
    df_display = df.reset_index()
    for ri, row_data in enumerate(df_display.itertuples(), start=4):
        ws_raw.row_dimensions[ri].height = 16
        shade = "FFF9F9F9" if ri % 2 == 0 else WHITE
        for ci, (field, label) in enumerate(raw_cols, 1):
            c = ws_raw.cell(row=ri, column=ci)
            if field == "year":
                val = row_data.Index if hasattr(row_data, 'Index') else row_data[0]
                val = int(val) if not pd.isna(val) else ""
            else:
                val = getattr(row_data, field, None)
                if val is not None and not pd.isna(val):
                    val = round(float(val), 2)
                else:
                    val = ""
            c.value = val
            c.font = cell_font(size=9)
            c.fill = fill(shade)
            c.border = thin_border()
            c.alignment = right_align() if ci > 1 else center()
            if field == "year":
                c.font = Font(name="Arial", bold=True, size=9)

    # Source notes at bottom
    last_row = 4 + len(df_display)
    ws_raw.row_dimensions[last_row + 1].height = 8
    note_cell = ws_raw.cell(row=last_row + 2, column=1)
    note_cell.value = ("SOURCES: Unemployment/LFPR/Income = FRED API | "
                       "Homicide = FBI UCR embedded | "
                       "Incarceration = BJS NPS embedded | "
                       "Overdose = CDC NCHS embedded")
    note_cell.font = Font(name="Arial", size=8, italic=True, color=GRAY)
    ws_raw.merge_cells(f"A{last_row+2}:H{last_row+2}")

    # ════════════════════════════════════════
    # SHEET 3: CLOCK ENGINE
    # ════════════════════════════════════════
    ws_eng = wb.create_sheet("⚙ Clock Engine")
    ws_eng.sheet_view.showGridLines = False

    eng_cols = [
        ("year", "Year", 8),
        ("ring_C_score", "Ring C\nBehavioral\n(Normalized)", 14),
        ("ring_B_score", "Ring B\nNormative\n(Placeholder)", 14),
        ("ring_A_score", "Ring A\nStructural\n(Normalized)", 14),
        ("clock_position", "Clock\nPosition\n(-1 to +1)", 14),
        ("clock_position_10pt", "Clock\nPosition\n(-10 to +10)", 14),
        ("velocity", "Velocity\n(3yr Δ)", 14),
        ("acceleration", "Acceleration\n(3yr Δ²)", 14),
        ("saddle_score", "Saddle\nScore\n(0-3)", 10),
        ("saddle_label", "Saddle Signal", 22),
        ("pressure_index", "Pressure\nIndex", 14),
    ]

    for ci, (field, label, width) in enumerate(eng_cols, 1):
        ws_eng.column_dimensions[get_column_letter(ci)].width = width

    ws_eng.row_dimensions[1].height = 10
    ws_eng.row_dimensions[2].height = 14

    # Title
    ws_eng.merge_cells("A3:K3")
    tc = ws_eng["A3"]
    tc.value = "HARM TOLERANCE CLOCK — PENDULUM DYNAMICS ENGINE"
    tc.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    tc.fill = fill(DARK_BG)
    tc.alignment = center()
    ws_eng.row_dimensions[3].height = 28

    # Subheader explanation
    ws_eng.merge_cells("A4:K4")
    sc = ws_eng["A4"]
    sc.value = ("Position = weighted Ring composite | "
                "Velocity = 3yr rate of change | "
                "Acceleration = rate of change of velocity | "
                "Saddle Score: 3 = STRONG predictive signal ⚡")
    sc.font = Font(name="Arial", size=9, italic=True, color=GRAY)
    sc.fill = fill(MID_BG)
    sc.alignment = center()
    ws_eng.row_dimensions[4].height = 18

    # Column headers
    for ci, (field, label, width) in enumerate(eng_cols, 1):
        c = ws_eng.cell(row=5, column=ci)
        c.value = label
        c.font = hdr_font(size=9)
        c.fill = fill(DARK_BG)
        c.alignment = center()
        c.border = thin_border()
    ws_eng.row_dimensions[5].height = 44

    # Data rows with conditional formatting
    for ri, row_data in enumerate(df_display.itertuples(), start=6):
        ws_eng.row_dimensions[ri].height = 16
        yr_val = int(row_data.Index) if not pd.isna(row_data.Index) else ""
        saddle = df.loc[row_data.Index, "saddle_score"] if row_data.Index in df.index else 0
        saddle_val = saddle if not pd.isna(saddle) else 0

        # Row background based on saddle strength
        if saddle_val == 3:
            row_bg = "FFFFF8DC"  # cornsilk — strong signal
        elif saddle_val == 2:
            row_bg = "FFFFF0F0"  # light pink — moderate
        else:
            row_bg = "FFF9F9F9" if ri % 2 == 0 else WHITE

        for ci, (field, label, width) in enumerate(eng_cols, 1):
            c = ws_eng.cell(row=ri, column=ci)
            if field == "year":
                c.value = yr_val
                c.font = Font(name="Arial", bold=True, size=9)
            elif field == "saddle_label":
                val = df.loc[row_data.Index, field] if row_data.Index in df.index else ""
                c.value = val
                if "STRONG" in str(val):
                    c.font = Font(name="Arial", bold=True, size=9, color="FF8B0000")
                elif "MODERATE" in str(val):
                    c.font = Font(name="Arial", bold=True, size=9, color="FFFF6600")
                else:
                    c.font = cell_font(size=9)
            else:
                val = df.loc[row_data.Index, field] if row_data.Index in df.index else None
                if val is not None and not pd.isna(val):
                    c.value = round(float(val), 4)
                else:
                    c.value = ""
                c.font = cell_font(size=9)

                # Color-code position column
                if field == "clock_position_10pt" and c.value != "":
                    pos = float(c.value)
                    if pos < -5:
                        c.font = Font(name="Arial", size=9, color="FFCC0000", bold=True)
                    elif pos < 0:
                        c.font = Font(name="Arial", size=9, color="FFFF6600")
                    elif pos > 5:
                        c.font = Font(name="Arial", size=9, color="FF006600", bold=True)
                    elif pos > 0:
                        c.font = Font(name="Arial", size=9, color="FF009900")

            c.fill = fill(row_bg)
            c.border = thin_border()
            c.alignment = center() if ci == 1 else right_align()

    # ════════════════════════════════════════
    # SHEET 4: SADDLE SIGNALS ONLY
    # ════════════════════════════════════════
    ws_saddle = wb.create_sheet("🔮 Saddle Signals")
    ws_saddle.sheet_view.showGridLines = False

    ws_saddle.column_dimensions["A"].width = 3
    ws_saddle.column_dimensions["B"].width = 10
    ws_saddle.column_dimensions["C"].width = 16
    ws_saddle.column_dimensions["D"].width = 14
    ws_saddle.column_dimensions["E"].width = 14
    ws_saddle.column_dimensions["F"].width = 14
    ws_saddle.column_dimensions["G"].width = 24
    ws_saddle.column_dimensions["H"].width = 50

    ws_saddle.row_dimensions[2].height = 36
    ws_saddle.merge_cells("B2:H2")
    c = ws_saddle["B2"]
    c.value = "🔮 HARM TOLERANCE CLOCK — SADDLE POINT DETECTION"
    c.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill = fill(DARK_BG)
    c.alignment = center()

    ws_saddle.row_dimensions[3].height = 16
    ws_saddle.merge_cells("B3:H3")
    c = ws_saddle["B3"]
    c.value = "SCORE 3 = STRONG SADDLE ⚡ | Velocity < 0 + Acceleration > 0 + Zero-crossing confirmed | These are your predictive turning point signals"
    c.font = Font(name="Arial", size=9, italic=True, color=GRAY)
    c.fill = fill(MID_BG)
    c.alignment = center()

    saddle_headers = ["Year", "Clock\nPosition", "Velocity", "Acceleration", "Saddle\nScore", "Signal", "Interpretation"]
    for ci, h in enumerate(saddle_headers, 2):
        ws_saddle.row_dimensions[4].height = 36
        c = ws_saddle.cell(row=4, column=ci)
        c.value = h
        c.font = hdr_font(size=10)
        c.fill = fill(DARK_BG)
        c.alignment = center()
        c.border = thin_border()

    # Filter to saddle signals only (score >= 2)
    saddle_rows = df[df["saddle_score"] >= 2].copy()

    if len(saddle_rows) == 0:
        ws_saddle["B5"].value = "No saddle signals detected in this data range — check Ring B placeholder data"
        ws_saddle["B5"].font = Font(name="Arial", size=10, italic=True, color=GRAY)
    else:
        for ri, (yr, row) in enumerate(saddle_rows.iterrows(), start=5):
            ws_saddle.row_dimensions[ri].height = 28
            score = int(row["saddle_score"])

            row_color = "FFFFF8DC" if score == 3 else "FFFFF0F0"

            vals = [
                yr,
                round(row["clock_position_10pt"], 2) if not pd.isna(row["clock_position_10pt"]) else "",
                round(row["velocity"], 4) if not pd.isna(row["velocity"]) else "",
                round(row["acceleration"], 4) if not pd.isna(row["acceleration"]) else "",
                score,
                row["saddle_label"],
            ]

            # Interpretation
            if score == 3:
                interp = (f"⚡ STRONG SIGNAL: Maximum momentum loading detected. "
                         f"Clock still moving negative but deceleration confirmed. "
                         f"Check leading indicators now — reversal pressure building.")
            else:
                interp = (f"Moderate inflection signal. Velocity negative, "
                         f"acceleration positive. Monitor for confirmation.")

            vals.append(interp)

            for ci, val in enumerate(vals, 2):
                c = ws_saddle.cell(row=ri, column=ci)
                c.value = val
                c.fill = fill(row_color)
                c.border = thin_border()
                c.alignment = Alignment(wrap_text=True, vertical="center",
                                       horizontal="center" if ci <= 7 else "left")
                if ci == 2:  # year
                    c.font = Font(name="Arial", bold=True, size=11, color="FF000080")
                elif ci == 7:  # signal label
                    if "STRONG" in str(val):
                        c.font = Font(name="Arial", bold=True, size=10, color="FF8B0000")
                    else:
                        c.font = Font(name="Arial", bold=True, size=10, color="FFFF6600")
                else:
                    c.font = cell_font(size=10)

    # ════════════════════════════════════════
    # SHEET 5: CHARTS
    # ════════════════════════════════════════
    ws_chart = wb.create_sheet("📈 Charts")
    ws_chart.sheet_view.showGridLines = False

    # Write slim data for charting
    ws_chart["A1"].value = "Year"
    ws_chart["B1"].value = "Clock Position (-10 to +10)"
    ws_chart["C1"].value = "Velocity"
    ws_chart["D1"].value = "Acceleration"

    chart_df = df[["clock_position_10pt", "velocity", "acceleration"]].dropna()
    for ri, (yr, row) in enumerate(chart_df.iterrows(), start=2):
        ws_chart.cell(row=ri, column=1).value = yr
        ws_chart.cell(row=ri, column=2).value = round(row["clock_position_10pt"], 3)
        ws_chart.cell(row=ri, column=3).value = round(row["velocity"], 4)
        ws_chart.cell(row=ri, column=4).value = round(row["acceleration"], 4)

    nrows = len(chart_df) + 1

    # Chart 1: Clock Position over time
    chart1 = LineChart()
    chart1.title = "Harm Tolerance Clock Position (1960–present)"
    chart1.style = 10
    chart1.y_axis.title = "Position (-10 = Maximum Harm)"
    chart1.x_axis.title = "Year"
    chart1.height = 12
    chart1.width = 24

    data1 = Reference(ws_chart, min_col=2, max_col=2, min_row=1, max_row=nrows)
    cats1 = Reference(ws_chart, min_col=1, min_row=2, max_row=nrows)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.series[0].graphicalProperties.line.solidFill = ACCENT.replace("FF", "")
    chart1.series[0].graphicalProperties.line.width = 20000
    ws_chart.add_chart(chart1, "F2")

    # Chart 2: Velocity + Acceleration
    chart2 = LineChart()
    chart2.title = "Velocity & Acceleration (Pendulum Dynamics)"
    chart2.style = 10
    chart2.y_axis.title = "Rate of Change"
    chart2.x_axis.title = "Year"
    chart2.height = 12
    chart2.width = 24

    data2 = Reference(ws_chart, min_col=3, max_col=4, min_row=1, max_row=nrows)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats1)
    chart2.series[0].graphicalProperties.line.solidFill = "4080FF"
    chart2.series[0].graphicalProperties.line.width = 18000
    chart2.series[1].graphicalProperties.line.solidFill = "FF4040"
    chart2.series[1].graphicalProperties.line.width = 15000
    ws_chart.add_chart(chart2, "F22")

    # Save workbook
    wb.save(xlsx_path)
    print(f"[EXPORT] Excel saved: {xlsx_path}")

except Exception as e:
    print(f"[ERROR] Excel build failed: {e}")
    import traceback
    traceback.print_exc()


# ─────────────────────────────────────────
# SECTION 7: PRINT SUMMARY
# ─────────────────────────────────────────

print("\n" + "=" * 60)
print("CEREBRO PHASE 1 — SUMMARY REPORT")
print("=" * 60)

print(f"\nYears covered: {YEARS_START}–{YEARS_END}")
print(f"Clock position range: {df['clock_position_10pt'].min():.2f} to {df['clock_position_10pt'].max():.2f}")

print("\n── SADDLE SIGNALS DETECTED ──")
saddle_events = df[df["saddle_score"] >= 2][["clock_position_10pt", "velocity", "acceleration", "saddle_label"]]
if len(saddle_events) == 0:
    print("  None detected (check Ring B placeholder data)")
else:
    for yr, row in saddle_events.iterrows():
        print(f"  {yr}: {row['saddle_label']} | "
              f"Pos={row['clock_position_10pt']:.2f} | "
              f"Vel={row['velocity']:.4f} | "
              f"Acc={row['acceleration']:.4f}")

print("\n── RECENT CLOCK READINGS (last 10 years) ──")
recent = df.tail(10)[["clock_position_10pt", "velocity", "acceleration", "saddle_score"]]
print(recent.to_string())

print("\n── RING B STATUS ──")
if pew_trust:
    print("  ✓ Ring B: Pew trust in government loaded (1958–present)")
    print("  To add more: GSS batch (gssdataexplorer.norc.org) + Gallup crime")
else:
    print("  ⚠ Ring B: Placeholder — run cerebro_data_gather.py for Pew trust")
print("  To add more Ring B data:")
print("  1. GSS variable COURTS: 'Courts deal harshly with criminals' 1972–present")
print("     → https://gssdataexplorer.norc.org/variables/297/vshow")
print("  2. GSS variable POLHITOK: 'Police use violence' 1973–present")
print("     → https://gssdataexplorer.norc.org/variables/340/vshow")
print("  3. Gallup Law/Order priority: request from Gallup Analytics")
print("     → https://www.gallup.com/analytics/213617/gallup-analytics.aspx")
print("  4. Pew Trust in Government: 1958–present composite")
print("     → https://www.pewresearch.org/politics/datasets/")

print("\n── FILES CREATED ──")
print(f"  {csv_path}")
if os.path.exists(xlsx_path):
    print(f"  {xlsx_path}")

print("\n── NEXT: PHASE 1 VALIDATION ──")
print("  Expected strong saddle signals (validate algorithm):")
print("  ✓ 1968-1972: Crime rising, reform pressure loading → Civil Rights/Warren Court response")
print("  ✓ 1991-1994: Crime peak, political loading → 1994 Crime Bill (saddle then overcorrection)")
print("  ✓ 2008-2012: Incarceration peak, reform pressure → 2015-2020 decarceration push")
print("  ✓ 2020-2021: Harm surge (OD + violent crime) → current predictive loading")
print("\nRun script again after loading Ring B data for full clock picture.")
print("=" * 60)
